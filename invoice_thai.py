import re
import os

from pypdf import PdfReader, PdfWriter

# import module to manage pdf file 
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

PDF_PATH = r'C:\Users\saich\Desktop\Day 2'
all_file = os.listdir(PDF_PATH)


pdf_list = []

for f in all_file:
    if f.startswith('IV') and f.endswith('.pdf'):
        pdf_list.append(f)

header = ['Date', 'Customer Name', 'Telephone Number', 'Email', 'Total']

date_data = []
name_data = []
tel_data = []
email_data = []
total_data = []
allDataTable = []

for pdf in pdf_list:
    reader = PdfReader(os.path.join(PDF_PATH, pdf))
    page = reader.pages[0]

    text = page.extract_text()

    # clean data

    text = text.strip().split('\n')

    # clear list that is null
    clean_text = []
    for t in text:
        if t != '' and t != ' ':
            replace_text = t.strip()
            clean_text.append(replace_text)


    thai_date_regex = r"\d{1,2}\s[ก-๙]\w+\s\d{4}"
    date_compile = re.compile(thai_date_regex)
    date_result = date_compile.search(clean_text[1])
    date_data.append(date_result.group())

    name_regex = r"[ก-๙]+\s[ก-๙]+"
    name_compile = re.compile(name_regex)
    name_result = name_compile.search(clean_text[2].replace(' า', 'ำ'))
    name_data.append(name_result.group())

    tel_regex = r"\d{3}(-)?\d{3}(-)?\d{4}"
    tel_compile = re.compile(tel_regex)
    tel_result = tel_compile.search(clean_text[4])
    tel_data.append(tel_result.group())

    email_regex = r"[\w\-._]+@[\w+]+\.[\w]{2,3}(.)?(th)?"
    email_compile = re.compile(email_regex)
    email_result = email_compile.search(clean_text[5].replace(' ', ''))
    email_data.append(email_result.group())

    total_regex = r"(-)?\d{1,3}(?:,\d{3})*(.)?\d{0,2}"
    total_compile = re.compile(total_regex)
    total_result = total_compile.search(clean_text[-2])
    total_data.append(total_result.group())

def saveToPDF():
    new_file = 'invoiceThai.pdf'
    writer = PdfWriter()
    writer.add_blank_page(width=8.27*72, height=11.7*72) # create pdf size "A4"
    writer.write(new_file)

    # Font
    pdfmetrics.registerFont(TTFont('F1', os.path.join(PDF_PATH, 'THSarabunNew Bold.ttf')))
    pdfmetrics.registerFont(TTFont('F2', os.path.join(PDF_PATH, 'THSarabunNew.ttf')))

    # create table
    allDataTable.append(header)
    for d in zip(date_data, name_data, tel_data, email_data, total_data):
        allDataTable.append(d)


    pdf = SimpleDocTemplate(new_file)
    table = Table(allDataTable, rowHeights=30)
    table.setStyle(TableStyle(
        [('GRID', (0,0), (-1,-1), 0.5, colors.black),
         ('FONTNAME', (0,0), (-1,0), 'F1'),
         ('ALIGN', (0,0), (-1,0), 'CENTER'),
         ('FONTNAME', (0,1), (-1,-1), 'F1'),
         ('VALIGN', (0,0), (-1,-1), 'TOP'),
         ('FONTSIZE',(0,0), (-1,-1), 18)]
    ))


    pdf.build([table])

saveToPDF()

def saveToExcel():
    global total_data
    wb = Workbook()
    ws = wb.active

    for h in range(len(header)):
        ws.cell(row=1, column=h+1, value=header[h])

    for c in ['A', 'B', 'c', 'D', 'E']:
        ws[f'{c}1'].font = Font(bold=True)
        ws[f'{c}1'].alignment = Alignment(horizontal='center')
        ws.column_dimensions[f'{c}'].width = 30

    total_data = [float(t.replace(',','')) for t in total_data]
    for d in zip(date_data, name_data, tel_data, email_data, total_data):
        ws.append(d)

    wb.save('invoice.xlsx')

saveToExcel()



